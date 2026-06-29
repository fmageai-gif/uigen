"""Tests for the flexible masterlist importer and the Auditor Name field."""

from __future__ import annotations

from openpyxl import Workbook

from eqms.core.models import Audit
from eqms.data import masterlist_import


def _make_messy_workbook(path):
    """A workbook mimicking a real masterlist: many sheets, data not first,
    header row not at the top, organisation-specific column names."""
    wb = Workbook()
    # First sheet is an unrelated summary (must NOT be picked).
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Region", "Headcount"])
    summary.append(["APAC", 700])

    # The real roster, with a title row above the header and HP-style columns.
    ms = wb.create_sheet("Masterlist")
    ms.append(["HP Mainstream Roster", None, None, None, None])
    ms.append(["Genesys Name", "HPI", "TL FULL NAME", "MNGR FULL NAME",
               "Internal LOB"])
    ms.append(["Ivy Serata", "HPI052258", "Estanislao, Mei", "Medel, Lorelei",
               "II Post Sales"])
    ms.append(["Mariter Mantaring", "HPI054790", "Tangaha, Nelissa",
               "Hernandez, Nilo", "One Print"])
    wb.save(path)


def test_import_detects_sheet_and_maps_columns(tmp_path):
    path = tmp_path / "HP_Masterlist.xlsx"
    _make_messy_workbook(path)

    result = masterlist_import.import_agents(path)

    assert result.sheet == "Masterlist"
    assert result.header_row == 1            # second row, not the title
    assert len(result.agents) == 2
    first = result.agents[0]
    assert first.agent_name == "Ivy Serata"
    assert first.agent_eid == "HPI052258"    # mapped from HPI
    assert first.team_leader == "Estanislao, Mei"
    assert first.operations_manager == "Medel, Lorelei"
    assert first.lob == "II Post Sales"
    # No TL/OM email columns present -> left blank, not an error.
    assert first.tl_email == ""


def test_import_rejects_workbook_without_agents(tmp_path):
    wb = Workbook()
    wb.active.append(["Foo", "Bar"])
    wb.active.append([1, 2])
    path = tmp_path / "junk.xlsx"
    wb.save(path)

    import pytest

    with pytest.raises(ValueError):
        masterlist_import.import_agents(path)


def test_explicit_override_forces_mapping(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Genesys Name", "HPI", "ORACLE"])
    ws.append(["Ivy Serata", "HPI052258", "101029441"])
    path = tmp_path / "ml.xlsx"
    wb.save(path)

    # Force the EID to come from ORACLE instead of the default HPI.
    result = masterlist_import.import_agents(path, overrides={"agent_eid": "ORACLE"})
    assert result.agents[0].agent_eid == "101029441"


def test_tl_om_emails_derived_by_eid_lookup(tmp_path):
    """When the masterlist lacks TL/OM email columns but has EID + EMAIL ADDRESS
    for everyone, the importer derives leader emails via EID lookup."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["EID", "Genesys Name", "EMAIL ADDRESS", "TL FULL NAME", "TL EID",
               "MNGR FULL NAME", "MNGR EID", "Internal LOB"])
    # A team leader and a manager as their own rows.
    ws.append(["500", "Boss Lady", "boss@x.com", "", "", "", "", "Mgmt"])
    ws.append(["900", "Big Boss", "bigboss@x.com", "", "", "", "", "Mgmt"])
    # An agent reporting to EID 500 (TL) and 900 (manager).
    ws.append(["101", "Ivy Serata", "ivy@x.com", "Boss Lady", "500",
               "Big Boss", "900", "II Post Sales"])
    path = tmp_path / "ml.xlsx"
    wb.save(path)

    result = masterlist_import.import_agents(path)
    agent = next(a for a in result.agents if a.agent_eid == "101")
    assert agent.agent_email == "ivy@x.com"   # agent's own email auto-populated
    assert agent.tl_email == "boss@x.com"
    assert agent.om_email == "bigboss@x.com"
    assert "tl_email" not in result.missing  # derived, so not reported missing


def test_tl_om_emails_derived_by_name_when_no_eid(tmp_path):
    """When leaders are referenced only by name (no TL/MNGR EID columns), emails
    are still resolved by matching the leader's full name to the EMAIL ADDRESS
    column of their own row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Masterlist"
    ws.append(["Genesys Name", "FULL NAME", "HPI", "EMAIL ADDRESS",
               "TL FULL NAME", "MNGR FULL NAME", "REGION", "WD LOB"])
    ws.append(["Boss Lady", "Estanislao, Mei", "HPI900", "boss@x.com",
               "", "", "AMS", "Mgmt"])
    ws.append(["Ivy", "Serata, Ivy", "HPI052", "ivy@x.com",
               "Estanislao, Mei", "Estanislao, Mei", "AMS", "II Post Sales"])
    path = tmp_path / "ml.xlsx"
    wb.save(path)

    result = masterlist_import.import_agents(path)
    ivy = next(a for a in result.agents if a.agent_name == "Ivy")
    assert ivy.region == "AMS"               # Region mapped from REGION
    assert ivy.lob == "II Post Sales"        # LOB mapped from WD LOB
    assert ivy.tl_email == "boss@x.com"      # resolved by TL full-name match
    assert ivy.om_email == "boss@x.com"


def test_auditor_name_roundtrips_through_excel_row():
    a = Audit(audit_id="A1", auditor_name="Jane Auditor", agent="X")
    restored = Audit.from_row(dict(zip(Audit.HEADERS, a.to_row())))
    assert restored.auditor_name == "Jane Auditor"
    assert "Auditor Name" in Audit.HEADERS
