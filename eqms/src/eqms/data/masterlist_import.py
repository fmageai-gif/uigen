"""Flexible masterlist importer.

Real-world masterlists are messy: the agent data may live in any of many sheets,
the header row is not always the first row, and column names vary between
organisations and exports. This module locates the correct sheet + header row by
scoring how well each candidate maps onto the fields the application needs, then
maps the columns using a configurable alias table.

The result is always normalised to the canonical :class:`~eqms.core.models.Agent`
layout, so the rest of the application is unaffected by the source file's shape.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..core.logging_config import get_logger
from ..core.models import Agent

_log = get_logger(__name__)

#: Field → ordered list of accepted source-header names (case/space-insensitive).
#: Earlier entries win, so put the most specific/preferred names first. Exact
#: (normalised) matches only — this is deliberate so that "EID" does not also
#: capture "TL EID" / "MNGR EID", which belong to other people.
COLUMN_ALIASES: dict[str, list[str]] = {
    "agent_name": [
        "agent name", "genesys name", "full name", "agent", "name",
        "employee name", "advisor name",
    ],
    "agent_eid": [
        "agent eid", "eid", "hpi", "employee id", "emp id", "worker id",
        "oracle", "new iex", "iex", "msp bpids", "msp bpid",
    ],
    "agent_email": [
        "agent email", "email address", "email", "e-mail", "work email",
        "email id",
    ],
    "team_leader": [
        "team leader", "tl full name", "tl name", "tl", "team lead",
        "supervisor", "immediate superior full name",
    ],
    "operations_manager": [
        "operations manager", "mngr full name", "manager", "mngr name",
        "ops manager", "operation manager", "om",
    ],
    "region": ["region", "queue", "wd position", "skill", "site location"],
    "lob": [
        "wd lob", "lob", "internal lob", "lob (msa)", "line of business",
        "lob code",
    ],
    "tl_email": [
        "tl email", "tl email address", "team leader email", "tl mail",
    ],
    "om_email": [
        "om email", "mngr email", "manager email", "operations manager email",
        "om mail",
    ],
}

#: Extra source columns used only *during* import (not stored on Agent): the
#: person's own email plus their TL's / manager's EID. These let us derive the
#: TL/OM email addresses by looking each leader up by EID in the same sheet,
#: even when the masterlist has no dedicated "TL Email"/"OM Email" columns.
LOOKUP_ALIASES: dict[str, list[str]] = {
    "email": ["email address", "email", "e-mail", "email id", "work email"],
    "full_name": ["full name", "fullname"],
    "tl_eid": ["tl eid", "team leader eid", "supervisor eid",
               "immediate superior eid"],
    "om_eid": ["mngr eid", "manager eid", "om eid", "operations manager eid"],
}

#: Fields without which a candidate header row is not a usable agent table.
REQUIRED_FIELDS = ("agent_name",)

#: How many leading rows of each sheet to consider as a possible header row.
_HEADER_SCAN_DEPTH = 12


@dataclass(slots=True)
class ImportResult:
    """Outcome of an import: the agents plus a human-readable mapping report."""

    agents: list[Agent]
    sheet: str
    header_row: int
    mapping: dict[str, str]   # field -> source column header used
    missing: list[str]        # canonical fields with no source column

    def summary(self) -> str:
        mapped = ", ".join(f"{k}←{v}" for k, v in self.mapping.items())
        miss = f"; unmapped: {', '.join(self.missing)}" if self.missing else ""
        return (f"{len(self.agents)} agents from sheet '{self.sheet}' "
                f"(row {self.header_row + 1}). Mapped: {mapped}{miss}")


def _norm(value: object) -> str:
    return ("" if value is None else str(value)).strip().lower()


def _map_headers_for(headers: list, alias_table: dict[str, list[str]],
                     overrides: dict[str, str] | None = None) -> dict[str, int]:
    """Return ``field -> column index`` for a header row using ``alias_table``.

    Matching is exact (normalised) so, e.g., the agent's ``EID`` column is never
    confused with ``TL EID``/``MNGR EID``. ``overrides`` maps a field to an
    explicit source header name (set by the admin) which wins over the aliases.
    """
    normalised = [_norm(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, aliases in alias_table.items():
        if overrides and field in overrides:
            override = _norm(overrides[field])
            if override in normalised:
                mapping[field] = normalised.index(override)
                continue
        for alias in aliases:
            if alias in normalised:
                mapping[field] = normalised.index(alias)
                break
    return mapping


def _map_headers(headers: list, overrides: dict[str, str] | None = None
                 ) -> dict[str, int]:
    """Map a header row onto the canonical :class:`Agent` fields."""
    return _map_headers_for(headers, COLUMN_ALIASES, overrides)


def _score(mapping: dict[str, int]) -> int:
    score = len(mapping)
    for required in REQUIRED_FIELDS:
        if required not in mapping:
            score -= 100
    # Reward having an identifier and a leadership column — hallmarks of a
    # genuine agent roster rather than an unrelated lookup sheet.
    if "agent_eid" in mapping:
        score += 1
    if "team_leader" in mapping or "operations_manager" in mapping:
        score += 1
    return score


def detect(path: Path, overrides: dict[str, str] | None = None):
    """Locate the best (sheet, header_row, mapping) in the workbook."""
    wb = load_workbook(path, read_only=True, data_only=True)
    best = None  # (score, sheet, header_row, mapping, headers)
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for hr, row in enumerate(ws.iter_rows(values_only=True)):
                if hr >= _HEADER_SCAN_DEPTH:
                    break
                headers = list(row)
                mapping = _map_headers(headers, overrides)
                score = _score(mapping)
                if best is None or score > best[0]:
                    best = (score, sheet, hr, mapping, headers)
    finally:
        wb.close()
    return best


def import_agents(path: str | Path, overrides: dict[str, str] | None = None
                  ) -> ImportResult:
    """Read ``path`` and return an :class:`ImportResult` of normalised agents.

    Raises
    ------
    ValueError
        If no sheet contains a recognisable agent table.
    """
    path = Path(path)
    best = detect(path, overrides)
    if best is None or best[0] < 0:
        raise ValueError(
            "Could not find an agent table in this workbook. Expected a sheet "
            "with at least a name column (e.g. 'Agent Name', 'Genesys Name' or "
            "'Full Name')."
        )

    _score_val, sheet, header_row, mapping, headers = best
    field_to_header = {f: str(headers[i]) for f, i in mapping.items()}
    lookups = _map_headers_for(headers, LOOKUP_ALIASES)

    wb = load_workbook(path, read_only=True, data_only=True)
    agents: list[Agent] = []
    leaders: list[tuple[Agent, str, str]] = []  # (agent, tl_eid, om_eid)
    email_by_eid: dict[str, str] = {}
    email_by_name: dict[str, str] = {}
    try:
        ws = wb[sheet]
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx <= header_row:
                continue
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            agent = _row_to_agent(row, mapping)
            if not (agent.agent_name or agent.agent_eid):
                continue
            agents.append(agent)
            # The email used to resolve THIS person when they are someone's TL/OM
            # comes from the EMAIL ADDRESS column (== the agent's own email).
            own_email = agent.agent_email or _cell(row, lookups.get("email"))
            full_name = _cell(row, lookups.get("full_name"))
            if own_email:
                if agent.agent_eid:
                    email_by_eid[_norm(agent.agent_eid)] = own_email
                # Index by every name form so a TL/OM referenced by full name or
                # genesys name resolves regardless of which the masterlist uses.
                for name in (full_name, agent.agent_name):
                    if name:
                        email_by_name[_norm(name)] = own_email
            leaders.append((
                agent,
                _cell(row, lookups.get("tl_eid")),
                _cell(row, lookups.get("om_eid")),
            ))
    finally:
        wb.close()

    # Second pass: resolve each agent's TL/OM email from the EMAIL ADDRESS
    # column of their leader's row — matched first by EID, then by name, so it
    # works whichever identifier the masterlist links leaders by.
    derived = 0
    for agent, tl_eid, om_eid in leaders:
        if not agent.tl_email:
            agent.tl_email = (
                email_by_eid.get(_norm(tl_eid), "")
                or email_by_name.get(_norm(agent.team_leader), "")
            )
            derived += 1 if agent.tl_email else 0
        if not agent.om_email:
            agent.om_email = (
                email_by_eid.get(_norm(om_eid), "")
                or email_by_name.get(_norm(agent.operations_manager), "")
            )

    missing = [f for f in COLUMN_ALIASES
               if f not in mapping and f not in ("tl_email", "om_email")]
    if "tl_email" not in mapping and not derived:
        missing.append("tl_email")

    result = ImportResult(agents=agents, sheet=sheet, header_row=header_row,
                          mapping=field_to_header, missing=missing)
    if derived:
        _log.info("Derived %d TL emails by EID lookup", derived)
    _log.info("Masterlist import: %s", result.summary())
    return result


def _cell(row, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    return "" if value is None else str(value).strip()


def _row_to_agent(row, mapping: dict[str, int]) -> Agent:
    return Agent(
        agent_name=_cell(row, mapping.get("agent_name")),
        agent_eid=_cell(row, mapping.get("agent_eid")),
        agent_email=_cell(row, mapping.get("agent_email")),
        team_leader=_cell(row, mapping.get("team_leader")),
        operations_manager=_cell(row, mapping.get("operations_manager")),
        region=_cell(row, mapping.get("region")),
        lob=_cell(row, mapping.get("lob")),
        tl_email=_cell(row, mapping.get("tl_email")),
        om_email=_cell(row, mapping.get("om_email")),
    )
