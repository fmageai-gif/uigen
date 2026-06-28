"""Low-level openpyxl read/write helpers operating on a local ``.xlsx`` path.

Both storage backends ultimately materialise a workbook on local disk (the
SharePoint backend downloads first) and then use these functions to read or
mutate it. Keeping the openpyxl details here means the backends only worry
about *transport* (local copy vs SharePoint round-trip), not Excel mechanics.

All values are read and written as strings to keep the data model simple and
predictable; numeric/date interpretation is the responsibility of the service
layer that knows each column's semantics.
"""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook

from ..core.logging_config import get_logger

_log = get_logger(__name__)


def create_workbook(path: Path, sheet_name: str, headers: Sequence[str]) -> None:
    """Create a new workbook at ``path`` with a single header row."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    _style_header(ws, len(headers))
    wb.save(path)
    _log.debug("Created workbook %s (sheet=%s)", path.name, sheet_name)


def ensure_sheet(path: Path, sheet_name: str, headers: Sequence[str]) -> None:
    """Ensure ``sheet_name`` exists in ``path`` with the given header row.

    Creates the workbook if missing, adds the sheet if absent, and writes the
    header row if the sheet is empty. Existing data is preserved.
    """
    if not path.exists():
        create_workbook(path, sheet_name, headers)
        return
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(list(headers))
        _style_header(ws, len(headers))
        wb.save(path)
    else:
        ws = wb[sheet_name]
        if ws.max_row == 0 or all(c.value is None for c in ws[1]):
            ws.delete_rows(1, ws.max_row or 1)
            ws.append(list(headers))
            _style_header(ws, len(headers))
            wb.save(path)


def read_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, str]]:
    """Read a sheet into a list of header→value dictionaries.

    The first row is treated as the header. Empty trailing rows are skipped.
    Returns an empty list if the workbook or sheet does not exist.
    """
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    except KeyError:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header]

    records: list[dict[str, str]] = []
    for row in rows_iter:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            value = row[idx] if idx < len(row) else None
            record[key] = "" if value is None else str(value).strip()
        records.append(record)
    wb.close()
    return records


def write_rows(
    path: Path,
    sheet_name: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
) -> None:
    """Replace the entire contents of ``sheet_name`` with ``headers`` + ``rows``.

    Other sheets in the workbook are left untouched.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        # Keep the target sheet first for readability.
        wb.move_sheet(ws, -(len(wb.sheetnames) - 1))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    ws.append(list(headers))
    _style_header(ws, len(headers))
    for row in rows:
        ws.append(list(row))
    _autosize(ws, len(headers))
    wb.save(path)
    _log.debug("Wrote %d rows to %s!%s", len(rows), path.name, sheet_name)


def append_row(
    path: Path,
    sheet_name: str,
    headers: Sequence[str],
    row: Sequence[str],
) -> None:
    """Append a single row, creating the sheet/header if necessary."""
    ensure_sheet(path, sheet_name, headers)
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.append(list(row))
    wb.save(path)


def _style_header(ws, ncols: int) -> None:
    """Apply a simple bold header style (HP-inspired blue)."""
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor="0F4C81")  # HP-style enterprise blue
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = "A2"


def _autosize(ws, ncols: int, max_width: int = 60) -> None:
    """Roughly auto-size columns based on content length."""
    from openpyxl.utils import get_column_letter

    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)
