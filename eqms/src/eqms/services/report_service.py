"""Reporting: on-demand and automatic monthly Excel reports.

Generates a formatted ``.xlsx`` report containing a summary sheet (KPIs and
breakdowns) and a detail sheet (the audits in scope), optionally embedding
charts rendered with matplotlib. Monthly reports run automatically when enabled
in settings and can also be produced on demand from the Reports view.
"""

from __future__ import annotations

import calendar
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from .. import config
from ..core.logging_config import get_logger
from ..core.models import Audit
from ..core.utils import parse_date, percentage, slugify
from ..data.audit_repository import AuditRepository
from ..data.settings_store import SettingsStore
from .dashboard_service import DashboardService

_log = get_logger(__name__)


@dataclass(slots=True)
class ReportResult:
    path: Path
    audit_count: int
    period_label: str


class ReportService:
    """Build Excel reports from the audit dataset."""

    def __init__(
        self,
        *,
        audits: AuditRepository | None = None,
        settings: SettingsStore | None = None,
        dashboard: DashboardService | None = None,
    ):
        self.audits = audits or AuditRepository()
        self.settings = settings or SettingsStore()
        self.dashboard = dashboard or DashboardService(self.audits)

    # -- public API ---------------------------------------------------------

    def generate_monthly(self, year: int, month: int,
                         output_dir: Path | None = None) -> ReportResult:
        """Generate a report for a specific month."""
        first = date(year, month, 1)
        last = date(year, month, calendar.monthrange(year, month)[1])
        label = first.strftime("%B %Y")
        audits = [a for a in self.audits.all() if _within(a, first, last)]
        path = self._build_workbook(audits, label, output_dir)
        return ReportResult(path=path, audit_count=len(audits), period_label=label)

    def generate_range(self, start: date, end: date,
                       output_dir: Path | None = None) -> ReportResult:
        """Generate a report for an arbitrary date range."""
        label = f"{start.isoformat()} to {end.isoformat()}"
        audits = [a for a in self.audits.all() if _within(a, start, end)]
        path = self._build_workbook(audits, label, output_dir)
        return ReportResult(path=path, audit_count=len(audits), period_label=label)

    def auto_monthly_due(self, today: date | None = None) -> tuple[int, int] | None:
        """Return the (year, month) to auto-generate today, or ``None``.

        Auto reports cover the *previous* month and run on the configured day of
        the current month.
        """
        if not self.settings.get_bool("report.auto_monthly", True):
            return None
        today = today or date.today()
        if today.day != self.settings.get_int("report.day_of_month", 1):
            return None
        prev_month = today.month - 1 or 12
        prev_year = today.year if today.month > 1 else today.year - 1
        return prev_year, prev_month

    # -- workbook construction ----------------------------------------------

    def _build_workbook(self, audits: list[Audit], label: str,
                        output_dir: Path | None) -> Path:
        output_dir = output_dir or (config.BACKUP_DIR / "reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / f"EQMS_Report_{slugify(label)}.xlsx"

        wb = Workbook()
        self._write_summary_sheet(wb.active, audits, label)
        self._write_detail_sheet(wb.create_sheet("Audit Detail"), audits)
        chart_png = self._render_chart(audits, output_dir, label)
        if chart_png:
            self._embed_chart(wb["Summary"], chart_png)
        wb.save(path)
        _log.info("Report generated: %s (%d audits)", path.name, len(audits))
        return path

    def _write_summary_sheet(self, ws, audits: list[Audit], label: str) -> None:
        from openpyxl.styles import Font

        ws.title = "Summary"
        kpis = self.dashboard.compute_kpis(audits, status="Report")
        ws["A1"] = f"{config.APP_SHORT_NAME} — Monthly Report"
        ws["A1"].font = Font(bold=True, size=14, color="0F4C81")
        ws["A2"] = f"Period: {label}"
        ws["A2"].font = Font(italic=True)

        valid = sum(1 for a in audits if not a.is_invalid and a.validation)
        invalid = sum(1 for a in audits if a.is_invalid)
        summary_rows = [
            ("Total Audits", len(audits)),
            ("Valid", valid),
            ("Invalid", invalid),
            ("Valid %", f"{percentage(valid, valid + invalid):.1f}%"),
            ("Invalid %", f"{percentage(invalid, valid + invalid):.1f}%"),
            ("Top QA", kpis.top_qa),
            ("Top Agent", kpis.top_agent),
            ("Top TL", kpis.top_tl),
            ("Top OM", kpis.top_om),
            ("Most Common Invalid Reason", kpis.common_invalid_reason),
        ]
        for i, (k, v) in enumerate(summary_rows, start=4):
            ws.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 40

    def _write_detail_sheet(self, ws, audits: list[Audit]) -> None:
        from openpyxl.styles import Font, PatternFill

        ws.append(list(Audit.HEADERS))
        fill = PatternFill("solid", fgColor="0F4C81")
        for col in range(1, len(Audit.HEADERS) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = fill
            c.font = Font(bold=True, color="FFFFFF")
        ws.freeze_panes = "A2"
        for a in audits:
            ws.append(a.to_row())

    def _render_chart(self, audits: list[Audit], output_dir: Path,
                      label: str) -> Path | None:
        """Render a validity breakdown bar chart to PNG (best-effort)."""
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except Exception as exc:  # noqa: BLE001
            _log.warning("matplotlib unavailable, skipping chart: %s", exc)
            return None

        breakdown = self.dashboard.validity_breakdown(audits)
        fig, ax = plt.subplots(figsize=(4, 3))
        ax.bar(list(breakdown.keys()), list(breakdown.values()),
               color=["#2f9e44", "#e03131"])
        ax.set_title(f"Validity Breakdown — {label}")
        ax.set_ylabel("Audits")
        fig.tight_layout()
        png = output_dir / f"chart_{slugify(label)}.png"
        fig.savefig(png, dpi=110)
        plt.close(fig)
        return png

    def _embed_chart(self, ws, png: Path) -> None:
        try:
            from openpyxl.drawing.image import Image as XLImage

            img = XLImage(str(png))
            ws.add_image(img, "D4")
        except Exception as exc:  # noqa: BLE001
            _log.warning("Could not embed chart: %s", exc)


def _within(a: Audit, start: date, end: date) -> bool:
    d = parse_date(a.created_at) or parse_date(a.date)
    return d is not None and start <= d <= end
